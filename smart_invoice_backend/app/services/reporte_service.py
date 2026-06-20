"""Servicio de generación de reportes administrativos (PDF, Excel, CSV)."""
from __future__ import annotations

import csv
from datetime import datetime
from decimal import Decimal

from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.core.exceptions import NotFoundError
from app.models.factura import Factura
from app.models.proveedor import Proveedor
from app.models.reporte import Reporte
from app.schemas.reporte import ReporteCreate
from app.services import bitacora_service

_EXT = {"PDF": "pdf", "EXCEL": "xlsx", "CSV": "csv"}


def get_report_data(db: Session, params: ReporteCreate) -> list[dict]:
    """Consulta las facturas según los filtros y devuelve filas planas."""
    stmt = (
        select(Factura)
        .options(
            joinedload(Factura.proveedor),
            joinedload(Factura.datos_extraidos),
        )
    )
    if params.fecha_inicio:
        stmt = stmt.where(Factura.created_at >= datetime.combine(params.fecha_inicio, datetime.min.time()))
    if params.fecha_fin:
        stmt = stmt.where(Factura.created_at <= datetime.combine(params.fecha_fin, datetime.max.time()))
    if params.proveedor_id is not None:
        stmt = stmt.where(Factura.proveedor_id == params.proveedor_id)
    if not params.incluir_rechazados:
        stmt = stmt.where(Factura.estado != "Rechazado")

    facturas = db.scalars(stmt.order_by(Factura.created_at.asc())).all()

    filas: list[dict] = []
    for f in facturas:
        d = f.datos_extraidos
        prov = f.proveedor
        filas.append({
            "id": f.id,
            "numero_factura": (d.numero_factura if d else None) or "—",
            "fecha": (d.fecha_factura.isoformat() if d and d.fecha_factura else "—"),
            "proveedor": (prov.nombre if prov else (d.nombre_proveedor_ocr if d else None)) or "—",
            "nit": (prov.nit if prov else (d.nit_ocr if d else None)) or "—",
            "subtotal": (d.subtotal if d else None) or Decimal("0"),
            "impuestos": (d.impuestos if d else None) or Decimal("0"),
            "total": (d.total if d else None) or Decimal("0"),
            "estado": f.estado,
        })
    return filas


def _resumen(filas: list[dict]) -> dict:
    total_monto = sum((r["total"] for r in filas), Decimal("0"))
    por_estado: dict[str, int] = {}
    for r in filas:
        por_estado[r["estado"]] = por_estado.get(r["estado"], 0) + 1
    return {
        "total_facturas": len(filas),
        "total_monto": total_monto,
        "por_estado": por_estado,
    }


def _build_filename(tipo: str) -> tuple[str, str]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"reporte_{timestamp}.{_EXT[tipo]}"
    ruta_relativa = f"{settings.REPORTS_DIR}/{nombre_archivo}"
    return nombre_archivo, ruta_relativa


def generate_csv(filas: list[dict], ruta_absoluta: str) -> None:
    columnas = ["id", "numero_factura", "fecha", "proveedor", "nit",
                "subtotal", "impuestos", "total", "estado"]
    with open(ruta_absoluta, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=columnas)
        writer.writeheader()
        for r in filas:
            writer.writerow({c: str(r[c]) for c in columnas})


def generate_excel(filas: list[dict], resumen: dict, ruta_absoluta: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    headers = ["ID", "No. Factura", "Fecha", "Proveedor", "NIT",
               "Subtotal", "Impuestos", "Total", "Estado"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="2E5A88")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for r in filas:
        ws.append([
            r["id"], r["numero_factura"], r["fecha"], r["proveedor"], r["nit"],
            float(r["subtotal"]), float(r["impuestos"]), float(r["total"]), r["estado"],
        ])

    ws.append([])
    ws.append(["RESUMEN"])
    ws.append(["Total de facturas", resumen["total_facturas"]])
    ws.append(["Monto total procesado", float(resumen["total_monto"])])
    for estado, cant in resumen["por_estado"].items():
        ws.append([f"Facturas {estado}", cant])

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 45)

    wb.save(ruta_absoluta)


def generate_pdf(filas: list[dict], resumen: dict, ruta_absoluta: str) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    doc = SimpleDocTemplate(
        ruta_absoluta, pagesize=landscape(letter),
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph("SmartInvoice — Reporte de Facturas", styles["Title"]),
        Paragraph(
            f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"],
        ),
        Spacer(1, 0.4 * cm),
        Paragraph(
            f"<b>Total de facturas:</b> {resumen['total_facturas']} &nbsp;&nbsp; "
            f"<b>Monto total procesado:</b> Q {resumen['total_monto']:,.2f}",
            styles["Normal"],
        ),
        Spacer(1, 0.5 * cm),
    ]

    data = [["ID", "No. Factura", "Fecha", "Proveedor", "NIT",
             "Subtotal", "Impuestos", "Total", "Estado"]]
    for r in filas:
        data.append([
            str(r["id"]), r["numero_factura"], r["fecha"], r["proveedor"][:28],
            r["nit"], f"{r['subtotal']:,.2f}", f"{r['impuestos']:,.2f}",
            f"{r['total']:,.2f}", r["estado"],
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5A88")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF3F8")]),
        ("ALIGN", (5, 1), (7, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabla)
    doc.build(elementos)


def generate(db: Session, params: ReporteCreate, usuario_id: int) -> Reporte:
    """Genera el reporte en el formato indicado y persiste su metadata."""
    filas = get_report_data(db, params)
    resumen = _resumen(filas)

    nombre_archivo, ruta_relativa = _build_filename(params.tipo)
    ruta_absoluta = str(settings.reports_path / nombre_archivo)

    if params.tipo == "CSV":
        generate_csv(filas, ruta_absoluta)
    elif params.tipo == "EXCEL":
        generate_excel(filas, resumen, ruta_absoluta)
    else:
        generate_pdf(filas, resumen, ruta_absoluta)

    nombre = params.nombre or f"Reporte {params.tipo} {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    reporte = Reporte(
        nombre=nombre,
        tipo=params.tipo,
        ruta_archivo=ruta_relativa,
        usuario_id=usuario_id,
        fecha_inicio=params.fecha_inicio,
        fecha_fin=params.fecha_fin,
        proveedor_id=params.proveedor_id,
        total_facturas=resumen["total_facturas"],
    )
    db.add(reporte)
    db.commit()
    db.refresh(reporte)

    bitacora_service.log(
        db, accion="GENERACION_REPORTE", estado="EXITOSO",
        usuario_id=usuario_id,
        resultado=f"Reporte '{nombre}' generado ({params.tipo}, "
                  f"{resumen['total_facturas']} facturas).",
    )
    return reporte


def get_by_id(db: Session, reporte_id: int) -> Reporte:
    reporte = db.get(Reporte, reporte_id)
    if not reporte:
        raise NotFoundError("Reporte no encontrado.")
    return reporte


def get_all(db: Session, *, skip: int = 0, limit: int = 50) -> tuple[int, list[Reporte]]:
    stmt = select(Reporte)
    total = len(list(db.scalars(stmt).all()))
    stmt = stmt.order_by(Reporte.created_at.desc()).offset(skip).limit(limit)
    return total, list(db.scalars(stmt).all())


def delete(db: Session, reporte_id: int) -> bool:
    from app.config import BASE_DIR

    reporte = get_by_id(db, reporte_id)
    ruta = BASE_DIR / reporte.ruta_archivo
    db.delete(reporte)
    db.commit()
    if ruta.exists():
        ruta.unlink()
    return True
